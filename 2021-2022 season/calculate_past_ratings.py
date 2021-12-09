import pandas as pd
import time
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter



#SCRAPE EACH POSITIONS RATINGS
class Gk():

    def __init__(self,page,wb):
        self.page=page
        self.wb=wb
    
    #read that csv
    #calculate their ratings
    def calculate_ratings(self):
        ws=self.wb.active
        print('-----CALCULATING GOALKEEPER AVERAGE RATINGS FOR ', str(self.page +1 ), 'MATCH------')
        time.sleep(5)
        df=pd.read_csv('D:\\Udemy\\personal data science projects\\Arsenal\\2021-2022 season\\arsenal_match_preds\\2021-2022 season\\'+str(self.page+1)+'_match.csv')
        df=df.drop(df.columns[[0]], axis=1)
        df.columns=['players','ratings']

        #cleaning through our data
        df['ratings']=df['ratings'].replace('-',0)
        df['ratings']=df['ratings'].apply(lambda x:(float(x)))

        #take only goalkeepers which play(ratings>0)
        goalkeeper=df[(df['players'].str.lower().str.contains('leno')&df['ratings']>0)|
        (df['players'].str.lower().str.contains('ramsdale')&df['ratings']>0)]
    
        print(goalkeeper)
        #for example both leno and ramsdale play that match, then we just find the average rating for the goalkeeper position
        count=len(goalkeeper)
        avg_gk_ratings=sum(goalkeeper['ratings'])/count
        print("AVERAGE GK RATINGS:",avg_gk_ratings)
        time.sleep(5)
        #FILL IN THE MATCHWEEK
        #FILL IN THE FIRST COLUMN
        column_char=get_column_letter(1)
        #EACH ROW EXCEPT FOR THE FIRST ROW
        ws[column_char+str(self.page+2)]=int(self.page)

        #FILL IN THE GK RATINGS
        gk_column=get_column_letter(2)
        #EXCEPT FOR THE FIRST ROW
        ws[gk_column+str(self.page+2)]=avg_gk_ratings
        self.wb.save(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx')
        time.sleep(5)

   
class Def(Gk):


    def calculate_ratings(self):
        ws=self.wb.active
        print('-----CALCULATING DEFENDER AVERAGE RATINGS FOR ', str(self.page +1 ), 'MATCH------')
        df=pd.read_csv('D:\\Udemy\\personal data science projects\\Arsenal\\2021-2022 season\\arsenal_match_preds\\2021-2022 season\\'+str(self.page+1)+'_match.csv')
        df=df.drop(df.columns[[0]], axis=1)
        df.columns=['players','ratings']

        #cleaning through our data
        df['ratings']=df['ratings'].replace('-',0)
        df['ratings']=df['ratings'].apply(lambda x:(float(x)))

    


        #take only defenders which play(ratings>0)
        defends=df[(df['players'].str.lower().str.contains('tierney')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('white')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('magal')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('holding')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('soares')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('tomiyasu')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('tavares')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('chambers')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('mari')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('kolasinac')& df['ratings']>0)]
        print(defends)
        #for example both leno and ramsdale play that match, then we just find the average rating for the goalkeeper position
        count=len(defends)
        avg_def_ratings=sum(defends['ratings'])/count
        print("AVERAGE DEFENDER RATINGS : " , avg_def_ratings)

        time.sleep(5)
        #FILL IN THE THIRD COLUMN
        def_column=get_column_letter(3)
        #EXCEPT FOR THE FIRST ROW
        ws[def_column+str(self.page+2)]=avg_def_ratings
        self.wb.save(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx')
        time.sleep(5)

class Mid(Gk):


     def calculate_ratings(self):
        ws=self.wb.active
        print('-----CALCULATING MIDFIELDERS AVERAGE RATINGS FOR ', str(self.page +1 ), 'MATCH------')
        df=pd.read_csv('D:\\Udemy\\personal data science projects\\Arsenal\\2021-2022 season\\arsenal_match_preds\\2021-2022 season\\'+str(self.page+1)+'_match.csv')
        df=df.drop(df.columns[[0]], axis=1)
        df.columns=['players','ratings']
        
        #cleaning through our data
        df['ratings']=df['ratings'].replace('-',0)
        df['ratings']=df['ratings'].apply(lambda x:(float(x)))

        #take only midfielders which play(ratings>0)
        midfields=df[(df['players'].str.lower().str.contains('partey')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('smith')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('saka')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('odegaard')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('maitland')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('lokonga')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('elneny')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('xhaka')& df['ratings']>0)]
        print(midfields)

        count=len(midfields)
        avg_mid_ratings=sum(midfields['ratings'])/count
        print('AVERAGE MIDFIELD RATINGS : ', avg_mid_ratings)

        time.sleep(5)
        #FILL IN THE 4TH COLUMN
        mid_column=get_column_letter(4)
        #EXCEPT FOR THE FIRST ROW
        ws[mid_column+str(self.page+2)]=avg_mid_ratings
        self.wb.save(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx')
        time.sleep(5)

class Attack(Gk):

    def calculate_ratings(self):
        ws=self.wb.active
        print('-----CALCULATING FORWARDS AVERAGE RATINGS FOR ', str(self.page +1 ), 'MATCH------')
        df=pd.read_csv('D:\\Udemy\\personal data science projects\\Arsenal\\2021-2022 season\\arsenal_match_preds\\2021-2022 season\\'+str(self.page+1)+'_match.csv')
        df=df.drop(df.columns[[0]], axis=1)
        df.columns=['players','ratings']

        #cleaning through our data
        df['ratings']=df['ratings'].replace('-',0)
        df['ratings']=df['ratings'].apply(lambda x:(float(x)))

        #take only forwards which play(ratings>0)
        forwards=df[(df['players'].str.lower().str.contains('lacazette')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('aubameyang')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('Pépé')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('balogun')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('nketiah')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('martinelli')& df['ratings']>0)|
        (df['players'].str.lower().str.contains('nelson')& df['ratings']>0)]
        print(forwards)


        count=len(forwards)
        avg_frwd_ratings=sum(forwards['ratings'])/count
        print('AVERAGE FORWARDS RATING : ', avg_frwd_ratings)

        time.sleep(5)
        #FILL IN THE 4TH COLUMN
        forw_column=get_column_letter(5)
        #EXCEPT FOR THE FIRST ROW
        ws[forw_column+str(self.page+2)]=avg_frwd_ratings
        self.wb.save(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx')
        time.sleep(5)

for i in range(0,15):
  
    goalkeeper=Gk(i,load_workbook(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx'))
    goalkeeper.calculate_ratings()
    

    defender=Def(i,load_workbook(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx'))
    defender.calculate_ratings()

    midfield=Mid(i,load_workbook(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx'))
    midfield.calculate_ratings()

    forwards=Attack(i,load_workbook(r'D:\Udemy\personal data science projects\Arsenal\2021-2022 season\arsenal_match_preds\2021-2022 season\past_ratings.xlsx'))
    forwards.calculate_ratings()
